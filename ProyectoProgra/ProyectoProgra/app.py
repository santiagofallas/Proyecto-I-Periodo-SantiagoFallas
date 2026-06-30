from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime
import json
import hashlib
import secrets

app = Flask(__name__)
# Generar una clave secreta segura
app.secret_key = secrets.token_hex(32)

EXCEL_FILE = 'contactos.xlsx'

# ==================== CONFIGURACIÓN DE USUARIOS ====================
# Usuarios registrados (email: contraseña)
# En un sistema real, las contraseñas deberían estar hasheadas
usuarios_registrados = {
    'Ejemplo@gmail.com': '12345678',
    'admin@ejemplo.com': 'admin123',
    'usuario@test.com': 'test123'
}

# ==================== FUNCIONES PARA EXCEL ====================
def crear_archivo_excel():
    """Crea el archivo Excel con los encabezados si no existe"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Apellido', 'Empresa', 'Cargo', 'Email', 
                   'Teléfono', 'Dirección', 'Cumpleaños', 'Notas', 'Estado', 
                   'Favorito', 'Fecha Creación', 'Última Interacción', 'Usuario Creador']
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2D5276", end_color="2D5276", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Ajustar anchos de columna
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(EXCEL_FILE)
        print(f"✅ Archivo {EXCEL_FILE} creado exitosamente")

def get_next_id():
    """Obtiene el siguiente ID disponible"""
    if not os.path.exists(EXCEL_FILE):
        return 1
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    
    wb.close()
    return max_id + 1

def cargar_contactos(usuario_actual=None):
    """Carga contactos del Excel (filtrados por usuario si se especifica)"""
    if not os.path.exists(EXCEL_FILE):
        crear_archivo_excel()
        return []
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        contactos = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0] is not None:  # Si tiene ID
                contacto = {
                    'id': int(row[0]) if row[0] else 0,
                    'nombre': str(row[1]) if row[1] else '',
                    'apellido': str(row[2]) if row[2] else '',
                    'empresa': str(row[3]) if row[3] else '',
                    'cargo': str(row[4]) if row[4] else '',
                    'email': str(row[5]) if row[5] else '',
                    'telefono': str(row[6]) if row[6] else '',
                    'direccion': str(row[7]) if row[7] else '',
                    'cumpleanos': str(row[8]) if row[8] else '',
                    'notas': str(row[9]) if row[9] else '',
                    'estado': str(row[10]) if row[10] else 'Activo',
                    'favorito': row[11] == 'True' or row[11] == True,
                    'fecha_creacion': str(row[12]) if row[12] else '',
                    'ultima_interaccion': str(row[13]) if row[13] else '',
                    'usuario_creador': str(row[14]) if row[14] else ''
                }
                # Filtrar por usuario actual si se especifica
                if not usuario_actual or contacto['usuario_creador'] == usuario_actual:
                    contactos.append(contacto)
        
        wb.close()
        return contactos
    except Exception as e:
        print(f"Error al cargar contactos: {e}")
        return []

def guardar_contacto(contacto):
    """Guarda un nuevo contacto en Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        next_row = ws.max_row + 1
        
        ws.cell(row=next_row, column=1, value=contacto['id'])
        ws.cell(row=next_row, column=2, value=contacto['nombre'])
        ws.cell(row=next_row, column=3, value=contacto['apellido'])
        ws.cell(row=next_row, column=4, value=contacto['empresa'])
        ws.cell(row=next_row, column=5, value=contacto['cargo'])
        ws.cell(row=next_row, column=6, value=contacto['email'])
        ws.cell(row=next_row, column=7, value=contacto['telefono'])
        ws.cell(row=next_row, column=8, value=contacto['direccion'])
        ws.cell(row=next_row, column=9, value=contacto['cumpleanos'])
        ws.cell(row=next_row, column=10, value=contacto['notas'])
        ws.cell(row=next_row, column=11, value=contacto['estado'])
        ws.cell(row=next_row, column=12, value=str(contacto['favorito']))
        ws.cell(row=next_row, column=13, value=contacto['fecha_creacion'])
        ws.cell(row=next_row, column=14, value=contacto['ultima_interaccion'])
        ws.cell(row=next_row, column=15, value=contacto['usuario_creador'])
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Error al guardar contacto: {e}")
        return False

def actualizar_contacto(contacto):
    """Actualiza un contacto existente en Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value == contacto['id']:
                row[1].value = contacto['nombre']
                row[2].value = contacto['apellido']
                row[3].value = contacto['empresa']
                row[4].value = contacto['cargo']
                row[5].value = contacto['email']
                row[6].value = contacto['telefono']
                row[7].value = contacto['direccion']
                row[8].value = contacto['cumpleanos']
                row[9].value = contacto['notas']
                row[10].value = contacto['estado']
                row[11].value = str(contacto['favorito'])
                row[13].value = contacto['ultima_interaccion']
                break
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Error al actualizar contacto: {e}")
        return False

def eliminar_contacto(id_contacto):
    """Elimina un contacto del Excel"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value == id_contacto:
                ws.delete_rows(row[0].row)
                break
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except Exception as e:
        print(f"Error al eliminar contacto: {e}")
        return False

# ==================== RUTAS DE FLASK ====================

@app.route('/', methods=['GET', 'POST'])
def login():
    """Página de login y procesamiento de autenticación"""
    # Si ya está logueado, redirigir al dashboard
    if 'usuario' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Recibir datos del formulario
        usuario_input = request.form.get('User')
        contrasena_input = request.form.get('Passw')
        
        # Verificar credenciales
        if usuario_input in usuarios_registrados and usuarios_registrados[usuario_input] == contrasena_input:
            # Guardar usuario en sesión
            session['usuario'] = usuario_input
            return jsonify({
                'success': True,
                'message': '¡Inicio de sesión exitoso!',
                'redirect': url_for('dashboard')
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Credenciales incorrectas. Prueba con Ejemplo@gmail.com / 12345678'
            }), 401
    
    # GET: mostrar página de login
    return render_template('login.html')

@app.route('/registro', methods=['POST'])
def registro():
    """Registro de nuevos usuarios"""
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    
    # Validaciones
    if not email or not password:
        return jsonify({'success': False, 'message': 'Datos incompletos'}), 400
    
    if email in usuarios_registrados:
        return jsonify({'success': False, 'message': 'El usuario ya existe'}), 409
    
    if len(password) < 4:
        return jsonify({'success': False, 'message': 'La contraseña debe tener al menos 4 caracteres'}), 400
    
    # Registrar nuevo usuario
    usuarios_registrados[email] = password
    return jsonify({'success': True, 'message': f'¡Registro exitoso! Bienvenido {email}. Ahora puedes iniciar sesión.'})

@app.route('/dashboard')
def dashboard():
    """Página principal después del login"""
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', usuario=session['usuario'])

@app.route('/api/contactos', methods=['GET'])
def get_contactos():
    """Obtener todos los contactos del usuario actual"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    contactos = cargar_contactos(session['usuario'])
    return jsonify(contactos)

@app.route('/api/contactos', methods=['POST'])
def add_contacto():
    """Agregar nuevo contacto"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    data = request.get_json()
    
    nuevo_contacto = {
        'id': get_next_id(),
        'nombre': data.get('nombre', ''),
        'apellido': data.get('apellido', ''),
        'empresa': data.get('empresa', ''),
        'cargo': data.get('cargo', ''),
        'email': data.get('email', ''),
        'telefono': data.get('telefono', ''),
        'direccion': data.get('direccion', ''),
        'cumpleanos': data.get('cumpleanos', ''),
        'notas': data.get('notas', ''),
        'estado': data.get('estado', 'Activo'),
        'favorito': data.get('favorito', False),
        'fecha_creacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'ultima_interaccion': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'usuario_creador': session['usuario']
    }
    
    if guardar_contacto(nuevo_contacto):
        return jsonify({'success': True, 'message': 'Contacto agregado', 'contacto': nuevo_contacto})
    else:
        return jsonify({'success': False, 'message': 'Error al guardar el contacto'}), 500

@app.route('/api/contactos/<int:id>', methods=['PUT'])
def update_contacto(id):
    """Actualizar contacto existente"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    data = request.get_json()
    contactos = cargar_contactos(session['usuario'])
    contacto = next((c for c in contactos if c['id'] == id), None)
    
    if not contacto:
        return jsonify({'error': 'Contacto no encontrado'}), 404
    
    # Actualizar solo los campos que vienen en la petición
    contacto['nombre'] = data.get('nombre', contacto['nombre'])
    contacto['apellido'] = data.get('apellido', contacto['apellido'])
    contacto['empresa'] = data.get('empresa', contacto['empresa'])
    contacto['cargo'] = data.get('cargo', contacto['cargo'])
    contacto['email'] = data.get('email', contacto['email'])
    contacto['telefono'] = data.get('telefono', contacto['telefono'])
    contacto['direccion'] = data.get('direccion', contacto['direccion'])
    contacto['cumpleanos'] = data.get('cumpleanos', contacto['cumpleanos'])
    contacto['notas'] = data.get('notas', contacto['notas'])
    contacto['estado'] = data.get('estado', contacto['estado'])
    contacto['favorito'] = data.get('favorito', contacto['favorito'])
    contacto['ultima_interaccion'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if actualizar_contacto(contacto):
        return jsonify({'success': True, 'message': 'Contacto actualizado'})
    else:
        return jsonify({'success': False, 'message': 'Error al actualizar'}), 500

@app.route('/api/contactos/<int:id>', methods=['DELETE'])
def delete_contacto(id):
    """Eliminar contacto"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    # Verificar que el contacto pertenece al usuario
    contactos = cargar_contactos(session['usuario'])
    contacto = next((c for c in contactos if c['id'] == id), None)
    
    if not contacto:
        return jsonify({'error': 'Contacto no encontrado'}), 404
    
    if eliminar_contacto(id):
        return jsonify({'success': True, 'message': 'Contacto eliminado'})
    else:
        return jsonify({'success': False, 'message': 'Error al eliminar'}), 500

@app.route('/api/contactos/toggle-favorite/<int:id>', methods=['PATCH'])
def toggle_favorite(id):
    """Cambiar estado de favorito"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    contactos = cargar_contactos(session['usuario'])
    contacto = next((c for c in contactos if c['id'] == id), None)
    
    if not contacto:
        return jsonify({'error': 'Contacto no encontrado'}), 404
    
    contacto['favorito'] = not contacto['favorito']
    contacto['ultima_interaccion'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if actualizar_contacto(contacto):
        return jsonify({'success': True, 'favorito': contacto['favorito']})
    else:
        return jsonify({'success': False, 'message': 'Error al actualizar'}), 500

@app.route('/api/descargar-reporte')
def descargar_reporte():
    """Descargar reporte Excel de contactos"""
    if 'usuario' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    # Crear un reporte solo con los contactos del usuario
    contactos_usuario = cargar_contactos(session['usuario'])
    
    # Crear nuevo Excel para el reporte
    wb = Workbook()
    ws = wb.active
    ws.title = f"Contactos_{session['usuario'].replace('@', '_')}"
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Apellido', 'Empresa', 'Cargo', 'Email', 
               'Teléfono', 'Dirección', 'Cumpleaños', 'Notas', 'Estado', 'Favorito']
    
    # Estilos para encabezados del reporte
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5276", end_color="2D5276", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    for row, contacto in enumerate(contactos_usuario, 2):
        ws.cell(row=row, column=1, value=contacto['id'])
        ws.cell(row=row, column=2, value=contacto['nombre'])
        ws.cell(row=row, column=3, value=contacto['apellido'])
        ws.cell(row=row, column=4, value=contacto['empresa'])
        ws.cell(row=row, column=5, value=contacto['cargo'])
        ws.cell(row=row, column=6, value=contacto['email'])
        ws.cell(row=row, column=7, value=contacto['telefono'])
        ws.cell(row=row, column=8, value=contacto['direccion'])
        ws.cell(row=row, column=9, value=contacto['cumpleanos'])
        ws.cell(row=row, column=10, value=contacto['notas'])
        ws.cell(row=row, column=11, value=contacto['estado'])
        ws.cell(row=row, column=12, value="Sí" if contacto['favorito'] else "No")
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar temporalmente y enviar
    reporte_path = f'reporte_{session["usuario"].replace("@", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(reporte_path)
    
    # Enviar archivo y luego eliminarlo
    from flask import after_this_request
    
    @after_this_request
    def remove_file(response):
        try:
            os.remove(reporte_path)
        except Exception as e:
            print(f"Error al eliminar archivo temporal: {e}")
        return response
    
    return send_file(reporte_path, as_attachment=True, download_name=f'contactos_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.pop('usuario', None)
    return redirect(url_for('login'))

# ==================== INICIALIZACIÓN ====================
if __name__ == '__main__':
    # Crear carpeta templates si no existe
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    # Crear carpeta static si no existe
    if not os.path.exists('static'):
        os.makedirs('static')
        os.makedirs('static/styles')
        os.makedirs('static/images')
    
    # Crear archivo Excel
    crear_archivo_excel()
    
    print("=" * 50)
    print("🚀 Servidor iniciado correctamente")
    print(f"📁 Archivo de contactos: {EXCEL_FILE}")
    print("📝 Credenciales de prueba:")
    print("   - Email: Ejemplo@gmail.com")
    print("   - Contraseña: 12345678")
    print("=" * 50)
    print("🌐 Abre tu navegador en: http://localhost:5000")
    print("=" * 50)
    
    app.run(debug=True, port=5000)